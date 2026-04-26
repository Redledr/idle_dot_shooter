#!/usr/bin/env python3
"""
cards_to_json.py
Converts CardDatabase.xlsx -> res://data/cards.json

Usage:
    python tools/cards_to_json.py
    python tools/cards_to_json.py --input path/to/CardDatabase.xlsx --output path/to/cards.json
"""

import json
import sys
import argparse
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

DEFAULT_INPUT  = Path(__file__).parent.parent / "CardDatabase.xlsx"
DEFAULT_OUTPUT = Path(__file__).parent.parent / "data" / "cards.json"

REQUIRED_COLUMNS = {"id", "name", "description", "rarity", "category", "effect_key", "effect_value"}
VALID_RARITIES   = {"common", "rare", "epic", "legendary"}
VALID_CATEGORIES = {"turret", "drone", "economy", "wildcard", "elemental", "chain", "dot", "nuke"}


def parse_args():
    parser = argparse.ArgumentParser(description="Convert CardDatabase.xlsx to cards.json")
    parser.add_argument("--input",  default=str(DEFAULT_INPUT),  help="Path to CardDatabase.xlsx")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT), help="Path to output cards.json")
    return parser.parse_args()


def convert(input_path: Path, output_path: Path) -> None:
    print(f"Reading: {input_path}")
    wb = openpyxl.load_workbook(input_path, data_only=True)
    ws = wb["Cards"]

    # Read header row
    headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]
    missing = REQUIRED_COLUMNS - set(headers)
    if missing:
        print(f"ERROR: Missing columns: {missing}")
        sys.exit(1)

    col = {name: idx for idx, name in enumerate(headers)}

    cards = []
    errors = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Skip empty rows
        if not row[col["id"]]:
            continue

        card_id      = str(row[col["id"]]).strip()
        name         = str(row[col["name"]]).strip()
        description  = str(row[col["description"]]).strip()
        rarity       = str(row[col["rarity"]]).strip().lower()
        category     = str(row[col["category"]]).strip().lower()
        effect_key   = str(row[col["effect_key"]]).strip() if row[col["effect_key"]] else ""
        effect_value = row[col["effect_value"]]
        notes        = str(row[col.get("notes", -1)]).strip() if col.get("notes") is not None and row[col.get("notes", -1)] else ""

        # Validate
        if rarity not in VALID_RARITIES:
            errors.append(f"Row {row_num}: invalid rarity '{rarity}' for card '{card_id}'")
        if category not in VALID_CATEGORIES:
            errors.append(f"Row {row_num}: invalid category '{category}' for card '{card_id}'")

        # Parse effect_value as number if possible
        try:
            effect_value = float(effect_value) if effect_value is not None else 0
            if effect_value == int(effect_value):
                effect_value = int(effect_value)
        except (TypeError, ValueError):
            effect_value = 0

        cards.append({
            "id":           card_id,
            "name":         name,
            "description":  description,
            "rarity":       rarity,
            "category":     category,
            "effect_key":   effect_key,
            "effect_value": effect_value,
        })

    if errors:
        print("VALIDATION ERRORS:")
        for e in errors:
            print(f"  {e}")
        sys.exit(1)

    # Check for duplicate IDs
    ids = [c["id"] for c in cards]
    dupes = [i for i in ids if ids.count(i) > 1]
    if dupes:
        print(f"ERROR: Duplicate card IDs: {list(set(dupes))}")
        sys.exit(1)

    # Write output
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump({"cards": cards}, f, indent=2, ensure_ascii=False)

    print(f"OK: Exported {len(cards)} cards -> {output_path}")
    for r in VALID_RARITIES:
        count = sum(1 for c in cards if c["rarity"] == r)
        print(f"  {r}: {count}")


if __name__ == "__main__":
    args = parse_args()
    convert(Path(args.input), Path(args.output))