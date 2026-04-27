#!/usr/bin/env python3
"""
cards_to_json.py
Converts CardDatabase.xlsx -> res://data/runtime/cards.json

Usage:
    python tools/python/cards_to_json.py
    python tools/python/cards_to_json.py --input path/to/CardDatabase.xlsx --output path/to/cards.json
"""

import json
import sys
import argparse
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_INPUT = PROJECT_ROOT / "data" / "source" / "CardDatabase.xlsx"
DEFAULT_OUTPUT = PROJECT_ROOT / "data" / "runtime" / "cards.json"
DEFAULT_EFFECT_SCRIPT = PROJECT_ROOT / "systems" / "progression" / "CardDatabase.gd"

REQUIRED_COLUMNS = {"id", "name", "description", "rarity", "category", "effect_key", "effect_value"}
VALID_RARITIES   = {"common", "rare", "epic", "legendary"}
VALID_CATEGORIES = {"turret", "drone", "economy", "wildcard", "elemental", "chain", "dot", "nuke"}

# Effect keys with direct logic in CardDatabase.apply_card().
IMPLEMENTED_EFFECT_KEYS = {
    "fire_rate_levels",
    "damage_levels",
    "balanced_boost",
    "damage_and_rate",
    "bullets_pierce",
    "fire_and_pierce",
    "drone_speed_levels",
    "drone_damage_levels",
    "drone_agility_levels",
    "drone_size_levels",
    "drone_cooldown_levels",
    "spawn_drone",
    "drone_combo",
    "drone_all",
    "orbs_per_kill",
    "pickup_radius_mul",
    "chain_pickup",
    "dot_value_levels",
    "spawn_count_levels",
    "orb_lifetime_add",
    "orb_gravity",
    "orb_nova",
    "orb_speed_mul",
    "orb_time_bonus",
    "auto_collect_radius",
    "economy_combo",
    "run_duration_add",
    "all_stats",
    "turret_all",
    "speed_combo",
    "rate_and_value",
    "time_and_dots",
    "density_boost",
    "efficient_trade",
    "random_stat",
    "random_three",
    "bullet_bounce",
    "bullet_wrap",
    "mirror_bullet",
    "chain_lightning",
    "chain_kill",
    "shockwave",
    "kill_stack",
    "kill_stack_cap",
    "execute_bonus",
    "volley_every",
    "frenzy_stack",
    "chain_orb_bonus",
    "orb_combo",
    "orb_frenzy",
    "dot_fire",
    "frost_slow",
    "frost_aoe",
    "frost_debuff",
    "hp_halve",
    "per_draw_bonus",
}

# Effect keys currently allowed in data, but they only flow into `main.card_flags`
# through CardDatabase's fallback branch until gameplay code consumes them.
FLAGGED_EFFECT_KEYS = {
    "acid_aoe",
    "acid_spread",
    "acid_stack",
    "black_hole_every",
    "bleed_orb_bonus",
    "bleed_stack",
    "cluster_every",
    "compound_bonus",
    "dot_acid",
    "dot_bleed",
    "dot_duration",
    "dot_fire_duration",
    "dot_fire_permanent",
    "dot_poison",
    "dot_spread",
    "emp_every",
    "extinction",
    "fire_bonus_full_hp",
    "fire_spread",
    "gravity_nuke",
    "nuke_every",
    "nuke_power",
    "orbital_every",
    "screen_nuke",
    "solar_flare",
    "void_pen",
}

VALID_EFFECT_KEYS = IMPLEMENTED_EFFECT_KEYS | FLAGGED_EFFECT_KEYS
FLOAT_EFFECT_KEYS = {
    "pickup_radius_mul",
    "orb_lifetime_add",
    "orb_speed_mul",
    "orb_time_bonus",
    "run_duration_add",
}


def parse_args():
    parser = argparse.ArgumentParser(description="Convert CardDatabase.xlsx to cards.json")
    parser.add_argument("--input",  default=str(DEFAULT_INPUT),  help="Path to CardDatabase.xlsx")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT), help="Path to output cards.json")
    parser.add_argument(
        "--validate-only",
        action="store_true",
        help="Validate the workbook and print a report without writing cards.json",
    )
    parser.add_argument(
        "--strict-effect-keys",
        action="store_true",
        help="Treat flagged future-system effect keys as validation errors instead of warnings",
    )
    return parser.parse_args()


def _parse_effect_value(value, row_num: int, card_id: str):
    if value is None or str(value).strip() == "":
        raise ValueError(f"Row {row_num}: missing effect_value for card '{card_id}'")

    try:
        parsed = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(
            f"Row {row_num}: invalid effect_value '{value}' for card '{card_id}'"
        ) from exc

    if parsed == int(parsed):
        return int(parsed)
    return parsed


def _validate_effect_key(effect_key: str, row_num: int, card_id: str, strict_effect_keys: bool, warnings: list[str], errors: list[str]) -> None:
    if not effect_key:
        errors.append(f"Row {row_num}: missing effect_key for card '{card_id}'")
        return

    if effect_key in IMPLEMENTED_EFFECT_KEYS:
        return

    if effect_key in FLAGGED_EFFECT_KEYS:
        message = (
            f"Row {row_num}: effect_key '{effect_key}' for card '{card_id}' "
            "is only flagged for future systems and is not directly implemented yet"
        )
        if strict_effect_keys:
            errors.append(message)
        else:
            warnings.append(message)
        return

    errors.append(
        f"Row {row_num}: unknown effect_key '{effect_key}' for card '{card_id}' "
        f"(not listed in {DEFAULT_EFFECT_SCRIPT.name} or the allowed flagged set)"
    )


def convert(input_path: Path, output_path: Path, validate_only: bool = False, strict_effect_keys: bool = False) -> None:
    print(f"Reading: {input_path}")
    wb = openpyxl.load_workbook(input_path, data_only=True)
    ws = wb["Cards"]

    # Read header row
    headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]
    missing = REQUIRED_COLUMNS - set(headers)
    if missing:
        print(f"ERROR: Missing columns: {missing}")
        sys.exit(1)

    col = {name: idx for idx, name in enumerate(headers)}

    cards = []
    errors = []
    warnings = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Skip empty rows
        if not row[col["id"]]:
            continue

        card_id      = str(row[col["id"]]).strip()
        name         = str(row[col["name"]]).strip()
        description  = str(row[col["description"]]).strip()
        rarity       = str(row[col["rarity"]]).strip().lower()
        category     = str(row[col["category"]]).strip().lower()
        effect_key   = str(row[col["effect_key"]]).strip() if row[col["effect_key"]] else ""
        effect_value = row[col["effect_value"]]
        notes        = str(row[col.get("notes", -1)]).strip() if col.get("notes") is not None and row[col.get("notes", -1)] else ""

        # Validate
        if rarity not in VALID_RARITIES:
            errors.append(f"Row {row_num}: invalid rarity '{rarity}' for card '{card_id}'")
        if category not in VALID_CATEGORIES:
            errors.append(f"Row {row_num}: invalid category '{category}' for card '{card_id}'")
        _validate_effect_key(effect_key, row_num, card_id, strict_effect_keys, warnings, errors)

        try:
            effect_value = _parse_effect_value(effect_value, row_num, card_id)
        except ValueError as exc:
            errors.append(str(exc))
            effect_value = 0

        if effect_key and effect_value != 0 and effect_key not in FLOAT_EFFECT_KEYS:
            if isinstance(effect_value, float) and not effect_value.is_integer():
                warnings.append(
                    f"Row {row_num}: effect_key '{effect_key}' for card '{card_id}' "
                    f"uses non-integer effect_value {effect_value}"
                )

        cards.append({
            "id":           card_id,
            "name":         name,
            "description":  description,
            "rarity":       rarity,
            "category":     category,
            "effect_key":   effect_key,
            "effect_value": effect_value,
        })

    if errors:
        print("VALIDATION ERRORS:")
        for e in errors:
            print(f"  {e}")
        sys.exit(1)

    # Check for duplicate IDs
    ids = [c["id"] for c in cards]
    dupes = [i for i in ids if ids.count(i) > 1]
    if dupes:
        print(f"ERROR: Duplicate card IDs: {list(set(dupes))}")
        sys.exit(1)

    if warnings:
        print("VALIDATION WARNINGS:")
        for warning in warnings:
            print(f"  {warning}")

    if validate_only:
        print(f"OK: Validated {len(cards)} cards")
        print(f"  implemented effect keys: {len(IMPLEMENTED_EFFECT_KEYS)}")
        print(f"  flagged effect keys: {len(FLAGGED_EFFECT_KEYS)}")
        print(f"  warnings: {len(warnings)}")
        return

    # Write output
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump({"cards": cards}, f, indent=2, ensure_ascii=False)

    print(f"OK: Exported {len(cards)} cards -> {output_path}")
    for r in VALID_RARITIES:
        count = sum(1 for c in cards if c["rarity"] == r)
        print(f"  {r}: {count}")


if __name__ == "__main__":
    args = parse_args()
    convert(
        Path(args.input),
        Path(args.output),
        validate_only=args.validate_only,
        strict_effect_keys=args.strict_effect_keys,
    )
